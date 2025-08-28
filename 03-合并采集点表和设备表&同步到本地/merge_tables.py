import json
import os
import pandas as pd
import pymysql
from sqlalchemy import create_engine, text
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment, PatternFill, Font
from openpyxl.utils import get_column_letter
from datetime import datetime


def apply_excel_formatting_all_sheets(filepath):
    """应用Excel格式设置到所有Sheet"""
    wb = load_workbook(filepath)

    # 定义样式
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    center_alignment = Alignment(horizontal='center', vertical='center')

    header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    header_font = Font(bold=True)

    # 遍历所有Sheet
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # 设置标题行样式（第一行）
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment

        # 设置行高
        ws.row_dimensions[1].height = 30
        for row_idx in range(2, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 16

        # 设置列宽和边框
        for col in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].width = 12

            for row in range(1, ws.max_row + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                cell.alignment = center_alignment

    wb.save(filepath)


def merge_generated_files(exported_files, merged_export_path, timestamp):
    """合并所有生成的Excel文件到一个文件中"""
    print("\n开始合并所有基地文件...")

    # 创建合并文件路径
    os.makedirs(merged_export_path, exist_ok=True)
    merged_filename = f"【合并】采集点+设备_{timestamp}.xlsx"
    merged_filepath = os.path.join(merged_export_path, merged_filename)

    # 创建一个新的Excel写入器
    with pd.ExcelWriter(merged_filepath, engine='openpyxl') as writer:
        # 初始化汇总数据
        all_data = []

        for file_info in exported_files:
            base = file_info['base']
            filepath = file_info['filepath']

            # 读取文件内容
            df = pd.read_excel(filepath)

            # 添加基地名称列（方便区分来源）
            df.insert(0, '基地', base)

            # 写入Excel，每个基地一个sheet
            sheet_name = f"{file_info['index']}_{base}"[:31]  # Excel sheet名最大31字符
            df.to_excel(writer, sheet_name=sheet_name, index=False)

            # 将数据添加到汇总列表
            all_data.append(df)

            print(f"已合并: {base}基地 -> {merged_filename}")

        # 创建汇总数据
        if all_data:
            summary_df = pd.concat(all_data, ignore_index=True)

            # 写入汇总sheet
            summary_sheet_name = "汇总"
            summary_df.to_excel(writer, sheet_name=summary_sheet_name, index=False)

            # 确保汇总sheet在第一个位置
            workbook = writer.book
            summary_sheet = workbook[summary_sheet_name]
            workbook.move_sheet(summary_sheet, offset=-len(workbook.sheetnames))

    # 格式化合并后的Excel文件的所有Sheet
    print("正在格式化合并后的文件...")
    apply_excel_formatting_all_sheets(merged_filepath)

    return merged_filepath


def merge_generated_files(exported_files, merged_export_path, timestamp):
    """合并所有生成的Excel文件到一个文件中"""
    print("\n开始合并所有基地文件...")

    # 创建合并文件路径
    os.makedirs(merged_export_path, exist_ok=True)
    merged_filename = f"【合并】采集点+设备_{timestamp}.xlsx"
    merged_filepath = os.path.join(merged_export_path, merged_filename)

    # 创建一个新的Excel写入器
    with pd.ExcelWriter(merged_filepath, engine='openpyxl') as writer:
        # 初始化汇总数据
        all_data = []

        for file_info in exported_files:
            base = file_info['base']
            filepath = file_info['filepath']

            # 读取文件内容
            df = pd.read_excel(filepath)

            # 添加基地名称列（方便区分来源）
            df.insert(0, '基地', base)

            # 写入Excel，每个基地一个sheet
            sheet_name = f"{file_info['index']}_{base}"[:31]  # Excel sheet名最大31字符
            df.to_excel(writer, sheet_name=sheet_name, index=False)

            # 将数据添加到汇总列表
            all_data.append(df)

            print(f"已合并: {base}基地 -> {merged_filename}")

        # 创建汇总数据
        if all_data:
            summary_df = pd.concat(all_data, ignore_index=True)

            # 写入汇总sheet
            summary_sheet_name = "汇总"
            summary_df.to_excel(writer, sheet_name=summary_sheet_name, index=False)

            # 确保汇总sheet在第一个位置
            workbook = writer.book
            summary_sheet = workbook[summary_sheet_name]
            workbook.move_sheet(summary_sheet, offset=-len(workbook.sheetnames))

    # 格式化合并后的Excel文件
    print("正在格式化合并后的文件...")
    apply_excel_formatting_all_sheets(merged_filepath)

    return merged_filepath


def merge_and_export():
    """主处理函数：合并表并导出Excel"""
    try:
        # 读取配置文件
        with open('config.json', 'r', encoding='utf-8') as f:
            config = json.load(f)

        # 从配置中获取所有必要信息
        db_config = config["db_config"]
        base_list = list(config["export_settings"].keys())
        column_mapping = config["column_mapping"]
        column_order = config["column_order"]
        export_settings = config["export_settings"]
        merged_export_path = config["merged_export_path"]  # 新添加的合并文件路径

        # 创建数据库连接
        engine = create_engine(
            f"mysql+pymysql://{db_config['user']}:{db_config['password']}@{db_config['host']}/{db_config['database']}?charset={db_config['charset']}"
        )

        # 生成统一的时间戳
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        exported_files = []  # 存储所有导出的文件信息

        # 处理每个基地
        for i, base in enumerate(base_list, 1):
            # 构建表名
            point_table = f"a采集点表_{i}_{base}"
            device_table = f"f设备表_{i}_{base}"
            merged_table = f"2_采集点+设备合并表_{i}_{base}"

            print(f"\n正在处理: {base}基地...")

            # 执行SQL合并表
            with engine.connect() as conn:
                # 删除已存在的合并表
                conn.execute(text(f"DROP TABLE IF EXISTS `{merged_table}`"))

                # 创建新合并表
                create_sql = text(f"""
                CREATE TABLE `{merged_table}` AS
                SELECT 
                    p.id AS point_id,
                    p.tag_name,
                    p.tag_code,
                    p.tag_desc,
                    p.ori_tag_name,
                    p.equipment_id,
                    p.general_attribute,
                    p.business_attribute,
                    p.classification,
                    p.verify_status,
                    d.id AS device_id,
                    d.equipment_name,
                    d.equipment_code,
                    d.base_name,
                    d.workshop,
                    d.workshop_section,
                    d.production_processes,
                    d.equipment_type,
                    d.equipment_sub_type,
                    d.equipment_attribute
                FROM `{point_table}` p
                LEFT JOIN `{device_table}` d ON p.equipment_id = d.id
                """)
                conn.execute(create_sql)

            # 读取合并表数据
            df = pd.read_sql_table(merged_table, engine)

            # 重命名列（英文列名 -> 中文列名）
            df.rename(columns=column_mapping, inplace=True)

            # 按配置重排列
            valid_columns = [col for col in column_order if col in df.columns]
            df = df[valid_columns]

            # 导出Excel
            export_path = export_settings.get(base, "")
            if not export_path:
                print(f"警告: {base}的导出路径未配置，跳过导出")
                continue

            os.makedirs(export_path, exist_ok=True)

            # 生成带时间戳的文件名
            filename = f"{merged_table}_{timestamp}.xlsx"
            filepath = os.path.join(export_path, filename)

            # 导出Excel
            df.to_excel(filepath, index=False)
            print(f"已导出: {filename}")

            # 应用Excel格式设置
            apply_excel_formatting_all_sheets(filepath)
            print(f"已格式化: {filename}")

            # 保存文件信息用于后续合并
            exported_files.append({
                'base': base,
                'index': i,
                'filepath': filepath,
                'filename': filename
            })

        print("\n所有基地处理完成！")

        # 合并所有生成的文件
        if exported_files:
            merged_filepath = merge_generated_files(exported_files, merged_export_path, timestamp)
            print(f"\n合并文件已生成: {merged_filepath}")
        else:
            print("没有文件可合并")

    except FileNotFoundError:
        print("错误: 未找到配置文件 config.json")
    except KeyError as e:
        print(f"配置错误: 缺少必要的配置项 - {str(e)}")
    except Exception as e:
        print(f"处理出错: {str(e)}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    merge_and_export()