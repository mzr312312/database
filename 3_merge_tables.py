import json
import os
import pandas as pd
import pymysql
import time
import openpyxl
from sqlalchemy import create_engine, text
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment, PatternFill, Font
from openpyxl.utils import get_column_letter
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor
from openpyxl.utils.dataframe import dataframe_to_rows


def apply_minimal_formatting(filepath):
    """仅设置基本格式（表头）"""
    wb = load_workbook(filepath)

    # 定义样式
    header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    header_font = Font(bold=True)
    center_alignment = Alignment(horizontal='center', vertical='center')

    # 遍历所有Sheet
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # 设置标题行样式（第一行）
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment

        # 设置行高
        ws.row_dimensions[1].height = 30

    wb.save(filepath)


def merge_generated_files(exported_files, merged_export_path, timestamp):
    """合并所有生成的Excel文件到一个文件中"""
    print("\n开始合并所有基地文件...")
    start_time = time.time()

    # 创建合并文件路径
    os.makedirs(merged_export_path, exist_ok=True)
    merged_filename = f"【合并】采集点+设备_{timestamp}.xlsx"
    merged_filepath = os.path.join(merged_export_path, merged_filename)

    # 创建一个新的Excel工作簿
    merged_wb = openpyxl.Workbook()

    # 创建汇总工作表
    summary_sheet = merged_wb.active
    summary_sheet.title = "汇总"

    # 初始化汇总数据列表
    all_data = []

    # 处理每个基地文件
    for file_info in exported_files:
        base = file_info['base']
        filepath = file_info['filepath']
        index = file_info['index']

        print(f"正在合并: {base}基地...")

        # 读取文件内容
        df = pd.read_excel(filepath)

        # 添加基地名称列（方便区分来源）
        df.insert(0, '基地', base)

        # 创建新工作表
        sheet_name = f"{index}_{base}"[:31]  # Excel sheet名最大31字符
        sheet = merged_wb.create_sheet(title=sheet_name)

        # 将数据写入工作表
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = sheet.cell(row=r_idx, column=c_idx, value=value)
                # 应用基本格式
                if r_idx == 1:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        # 将数据添加到汇总列表
        all_data.append(df)

    # 创建汇总数据
    if all_data:
        summary_df = pd.concat(all_data, ignore_index=True)

        # 将汇总数据写入汇总工作表
        for r_idx, row in enumerate(dataframe_to_rows(summary_df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = summary_sheet.cell(row=r_idx, column=c_idx, value=value)
                # 应用基本格式
                if r_idx == 1:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # 保存合并文件
    merged_wb.save(merged_filepath)
    print(f"合并完成! 耗时: {time.time() - start_time:.2f}秒")

    # 应用基本格式（仅表头）
    print("正在应用基本格式...")
    apply_minimal_formatting(merged_filepath)

    return merged_filepath


def process_base(i, base, config, timestamp):
    """处理单个基地的函数，用于并行处理"""
    try:
        # 从配置中获取必要信息
        db_config = config["db_config"]
        column_mapping = config["column_mapping"]
        column_order = config["column_order"]
        export_settings = config["export_settings"]

        # 创建数据库连接（每个线程需要自己的连接）
        engine = create_engine(
            f"mysql+pymysql://{db_config['user']}:{db_config['password']}@{db_config['host']}/{db_config['database']}?charset={db_config['charset']}"
        )

        print(f"正在处理: {base}基地...")

        # 构建表名
        point_table = f"a采集点表_{i}_{base}"
        device_table = f"f设备表_{i}_{base}"
        merged_table = f"2_采集点+设备合并表_{i}_{base}"

        # 执行SQL合并表
        with engine.connect() as conn:
            # 删除已存在的合并表
            print(f"  清理旧表: {merged_table}")
            conn.execute(text(f"DROP TABLE IF EXISTS `{merged_table}`"))

            # 创建新合并表
            print(f"  创建新表: {merged_table}")
            create_sql = text(f"""
            CREATE TABLE `{merged_table}` AS
            SELECT 
                p.id AS point_id,
                p.tag_name,
                p.tag_code,
                p.tag_desc,
                p.ori_tag_name,
                p.equipment_id,
                p.general_attribute,
                p.business_attribute,
                p.classification,
                p.verify_status,
                d.id AS device_id,
                d.equipment_name,
                d.equipment_code,
                d.base_name,
                d.workshop,
                d.workshop_section,
                d.production_processes,
                d.equipment_type,
                d.equipment_sub_type,
                d.equipment_attribute
            FROM `{point_table}` p
            LEFT JOIN `{device_table}` d ON p.equipment_id = d.id
            """)
            conn.execute(create_sql)

        # 读取合并表数据
        print(f"  读取数据: {merged_table}")
        df = pd.read_sql_table(merged_table, engine)

        # 重命名列（英文列名 -> 中文列名）
        print("  重命名列")
        df.rename(columns=column_mapping, inplace=True)

        # 按配置重排列
        print("  调整列顺序")
        valid_columns = [col for col in column_order if col in df.columns]
        df = df[valid_columns]

        # 导出Excel
        export_path = export_settings.get(base, "")
        if not export_path:
            print(f"警告: {base}的导出路径未配置，跳过导出")
            return None

        os.makedirs(export_path, exist_ok=True)

        # 生成带时间戳的文件名
        filename = f"{merged_table}_{timestamp}.xlsx"
        filepath = os.path.join(export_path, filename)

        # 导出Excel
        print(f"  导出到: {filepath}")
        df.to_excel(filepath, index=False)
        print(f"已导出: {filename}")

        # 应用基本格式（仅表头）
        print("  应用基本格式...")
        apply_minimal_formatting(filepath)

        return {
            'base': base,
            'index': i,
            'filepath': filepath,
            'filename': filename
        }

    except Exception as e:
        print(f"处理{base}基地时出错: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

    finally:
        # 确保释放数据库连接
        if 'engine' in locals():
            engine.dispose()


def merge_and_export():
    """主处理函数：合并表并导出Excel"""
    try:
        # 读取配置文件
        with open('3_config.json', 'r', encoding='utf-8') as f:
            config = json.load(f)

        # 从配置中获取所有必要信息
        base_list = list(config["export_settings"].keys())
        merged_export_path = config["merged_export_path"]

        # 生成统一的时间戳
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        exported_files = []  # 存储所有导出的文件信息

        print(f"开始处理 {len(base_list)} 个基地...")
        start_time = time.time()

        # 使用线程池并行处理
        with ThreadPoolExecutor(max_workers=min(4, len(base_list))) as executor:
            # 准备任务参数
            tasks = [(i + 1, base, config, timestamp) for i, base in enumerate(base_list)]

            # 提交所有任务
            futures = [executor.submit(process_base, *task) for task in tasks]

            # 收集结果
            for future in futures:
                result = future.result()
                if result:
                    exported_files.append(result)

        print(f"\n所有基地处理完成! 总耗时: {time.time() - start_time:.2f}秒")

        # 合并所有生成的文件
        if exported_files:
            merged_filepath = merge_generated_files(exported_files, merged_export_path, timestamp)
            print(f"\n合并文件已生成: {merged_filepath}")
        else:
            print("没有文件可合并")

    except FileNotFoundError:
        print("错误: 未找到配置文件 config.json")
    except KeyError as e:
        print(f"配置错误: 缺少必要的配置项 - {str(e)}")
    except Exception as e:
        print(f"处理出错: {str(e)}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    merge_and_export()
    print("程序已正常退出")